#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
受渡伝票生成模块 - 将填写了箱号的配分表转换为受渡伝票格式
"""
import openpyxl
from openpyxl import load_workbook
import xlrd
from typing import Dict, List, Optional
import os
from config import TemplateConfig, DeliveryNoteConfig

class DeliveryNoteGenerator:
    """受渡伝票生成器"""
    
    def __init__(self, input_path: str, template_path: str, output_path: str, start_no: int = None, prefix: str = "81"):
        """
        初始化生成器
        
        Args:
            input_path: 输入文件路径（填写了箱号的配分表）
            template_path: 模板文件路径（受渡伝票模板）
            output_path: 输出文件路径
            start_no: 起始编号 (可选)
            prefix: 受渡伝票NO前缀 (默认 "81")
        """
        self.input_path = input_path
        self.template_path = template_path
        self.output_path = output_path
        self.start_no = start_no
        self.prefix = prefix
        self.data_rows = []

    def process(self):
        """执行转换流程"""
        # 1. 读取输入文件数据
        self._read_input_data()
        
        # 2. 写入到输出模板
        self._write_to_template()
        
        return self.output_path

    def _read_input_data(self):
        """读取输入文件中的数据"""
        print(f"Reading input file: {self.input_path}")
        
        # 根据文件扩展名选择读取方式
        ext = os.path.splitext(self.input_path)[1].lower()
        
        if ext in ['.xlsx', '.xlsm']:
            self._read_xlsx()
        elif ext == '.xls':
            self._read_xls()
        else:
            raise ValueError(f"Unsupported file format: {ext}")
            
    def _read_xlsx(self):
        """使用openpyxl读取xlsx文件"""
        wb = load_workbook(self.input_path, data_only=True)
        
        # 遍历所有PT sheet
        for sheet_name in wb.sheetnames:
            # 假设PT sheet名称包含 "PT-"
            if "PT-" not in sheet_name:
                continue
                
            ws = wb[sheet_name]
            self._process_sheet_data_openpyxl(ws)

    def _process_sheet_data_openpyxl(self, ws):
        """处理单个sheet的数据 (openpyxl)"""
        # 1. 提取元数据 (Kanri No -> Brand)
        # E1 cell
        kanri_no = str(ws['E1'].value).strip() if ws['E1'].value else ""
        brand = kanri_no[:3] if len(kanri_no) >= 3 else kanri_no
        
        # 2. 提取SKU信息 (Row 2, 3, 4, Cols from I/8)
        # Row 2: Product Code (F2 is label, values start from I2)
        # Row 3: Color (F3 is label, values start from I3)
        # Row 4: Size (F4 is label, values start from I4)
        
        sku_map = {} # col_index -> {product_code, color, size}
        
        # Determine how many columns to read
        # Assuming data starts at col index 8 (Column I)
        current_col = 8
        while True:
            # Check if there is data in header rows
            prod_code = ws.cell(row=2, column=current_col+1).value
            if not prod_code:
                break
                
            color = ws.cell(row=3, column=current_col+1).value
            size = ws.cell(row=4, column=current_col+1).value
            
            sku_map[current_col] = {
                'product_code': str(prod_code),
                'color': str(color) if color is not None else "",
                'size': str(size) if size is not None else ""
            }
            current_col += 1
            
        # 3. 遍历数据行 (从 Row 6 / index 5 开始)
        # Col D (3): Store Code
        # Col F (5): CTN_NO
        # Col I (8) ~: SKU Quantities
        
        row_idx = 5
        # Removed auto-increment current_no
        # current_no = self.start_no if self.start_no is not None else None

        while row_idx < ws.max_row:
            row_num = row_idx + 1
            store_code = ws.cell(row=row_num, column=4).value # Col D
            ctn_no_raw = ws.cell(row=row_num, column=6).value # Col F
            
            # 停止条件：如果store_code为空，假设数据结束
            if not store_code:
                # 尝试检查下一行，或者直接结束
                # 这里简单判断：如果是空行则跳过或结束
                # 但配分表中间通常没有空行，直到合计行
                # 检查是否是合计行
                first_col_val = ws.cell(row=row_num, column=1).value
                if first_col_val and "合计" in str(first_col_val):
                    break
                if not first_col_val and not store_code:
                    break
                    
                row_idx += 1
                continue

            # 必须有 CTN_NO 才处理 (作为行有效的标志)
            if ctn_no_raw:
                try:
                    ctn_no_int = int(ctn_no_raw)
                    
                    # Logic update: Use start_no as offset if provided
                    if self.start_no is not None:
                         # 逻辑：seq = start_no + (ctn_no - 1)
                         # 假设 ctn_no 从 1 开始
                         seq_no = self.start_no + (ctn_no_int - 1)
                         slip_no = f"{self.prefix}{seq_no:04d}"
                    else:
                         slip_no = f"{self.prefix}{ctn_no_int:04d}"
                    
                    # 遍历SKU列，提取数量 > 0 的记录
                    for col_idx, sku_info in sku_map.items():
                        qty = ws.cell(row=row_num, column=col_idx+1).value
                        if qty and isinstance(qty, (int, float)) and qty > 0:
                            # 添加一条记录
                            self.data_rows.append({
                                'slip_no': slip_no,
                                'brand': brand,
                                'store_code': str(store_code),
                                'product_code': sku_info['product_code'],
                                'size': sku_info['size'],
                                'color': sku_info['color'],
                                'qty': int(qty)
                            })
                except ValueError:
                    print(f"Warning: Invalid CTN_NO '{ctn_no_raw}' at row {row_num}")
            
            row_idx += 1

    def _read_xls(self):
        """使用xlrd读取xls文件"""
        wb = xlrd.open_workbook(self.input_path)
        
        for sheet in wb.sheets():
            if "PT-" not in sheet.name:
                continue
            self._process_sheet_data_xlrd(sheet)

    def _process_sheet_data_xlrd(self, sheet):
        """处理单个sheet的数据 (xlrd)"""
        # 1. 提取元数据
        try:
            kanri_no = str(sheet.cell_value(0, 4)).strip() # E1 -> (0, 4)
            brand = kanri_no[:3] if len(kanri_no) >= 3 else kanri_no
        except:
            brand = ""
            
        # 2. 提取SKU信息
        sku_map = {}
        current_col = 8 # Col I
        
        while current_col < sheet.ncols:
            try:
                prod_code = sheet.cell_value(1, current_col) # Row 2
                if not prod_code:
                    break
                
                color = sheet.cell_value(2, current_col) # Row 3
                size = sheet.cell_value(3, current_col) # Row 4
                
                # xlrd 读取数字可能是 float
                if isinstance(color, float): color = str(int(color))
                if isinstance(size, float): size = str(int(size))
                if isinstance(prod_code, float): prod_code = str(int(prod_code))
                
                sku_map[current_col] = {
                    'product_code': str(prod_code),
                    'color': str(color),
                    'size': str(size)
                }
                current_col += 1
            except IndexError:
                break
                
        # 3. 遍历数据
        row_idx = 5 # Row 6
        # Removed auto-increment current_no
        # current_no = self.start_no if self.start_no is not None else None

        while row_idx < sheet.nrows:
            try:
                store_code = sheet.cell_value(row_idx, 3) # Col D
                ctn_no_raw = sheet.cell_value(row_idx, 5) # Col F
                
                if not store_code:
                    # Check for total row or end
                    first_val = sheet.cell_value(row_idx, 0)
                    if "合计" in str(first_val):
                        break
                    row_idx += 1
                    continue
                    
                if ctn_no_raw:
                    try:
                        ctn_no_val = 0
                        if isinstance(ctn_no_raw, (int, float)):
                            ctn_no_val = int(ctn_no_raw)
                        elif isinstance(ctn_no_raw, str) and ctn_no_raw.strip().isdigit():
                            ctn_no_val = int(ctn_no_raw)
                            
                        if ctn_no_val > 0:
                            if self.start_no is not None:
                                seq_no = self.start_no + (ctn_no_val - 1)
                                slip_no = f"{self.prefix}{seq_no:04d}"
                            else:
                                slip_no = f"{self.prefix}{ctn_no_val:04d}"
                            
                            for col_idx, sku_info in sku_map.items():
                                qty = sheet.cell_value(row_idx, col_idx)
                                if qty and isinstance(qty, (int, float)) and qty > 0:
                                    if isinstance(store_code, float): store_code = str(int(store_code))
                                    
                                    self.data_rows.append({
                                        'slip_no': slip_no,
                                        'brand': brand,
                                        'store_code': str(store_code),
                                        'product_code': sku_info['product_code'],
                                        'size': sku_info['size'],
                                        'color': sku_info['color'],
                                        'qty': int(qty)
                                    })
                    except ValueError:
                        pass
            except IndexError:
                break
            
            row_idx += 1

    def _write_to_template(self):
        """写入数据到输出模板"""
        print(f"Writing to template: {self.template_path}")
        
        # 必须使用 openpyxl 写入，所以模板必须是 xlsx
        # 如果模板是 xls，需要先转换或者报错
        # 这里的模板是 "③受渡伝票_模板（上传系统资料）.xls"
        # openpyxl 不支持 .xls。
        # 解决方案：
        # 1. 要求用户提供 xlsx 模板
        # 2. 或者我们尝试用 xlrd 读取 .xls 模板，然后用 openpyxl 创建新的 xlsx
        # 鉴于这是一个 "TemplateWriter"，保留样式很重要。
        # 如果模板是 .xls，我们只能用 xlutils (基于 xlwt/xlrd) 但 xlwt 不支持 xlsx 且对样式支持有限
        # 最好的方式：假设模板已被转换为 xlsx，或者我们在此处进行一次转换
        
        # 检查模板扩展名
        if self.template_path.lower().endswith('.xls'):
            # 这是一个问题，openpyxl 不能编辑 .xls
            # 我们需要提示用户或者自动转换
            # 为了简单起见，如果提供了 .xls，我们尝试寻找同名的 .xlsx
            xlsx_path = self.template_path + "x"
            if os.path.exists(xlsx_path):
                self.template_path = xlsx_path
            else:
                 # 这是一个临时的 workaround，实际上应该要求 .xlsx
                 # 或者我们可以尝试 "升级" 模板
                 raise ValueError("Template must be .xlsx format for writing. Please save the template as .xlsx.")

        wb = load_workbook(self.template_path)
        ws = wb.active # 假设写入第一个sheet
        
        # 开始写入
        start_row = DeliveryNoteConfig.WRITE_START_ROW + 1 # 1-indexed
        
        # 按照 B8 开始
        # B -> Col 2
        
        for idx, row_data in enumerate(self.data_rows):
            current_row = start_row + idx
            
            # No (序号) - Removed per request
            # ws.cell(row=current_row, column=DeliveryNoteConfig.COL_INDEX_NO + 1).value = idx + 1
            
            # 受渡伝票NO
            ws.cell(row=current_row, column=DeliveryNoteConfig.COL_INDEX_SLIP_NO + 1).value = row_data['slip_no']
            
            # ブランド
            ws.cell(row=current_row, column=DeliveryNoteConfig.COL_INDEX_BRAND + 1).value = row_data['brand']
            
            # 店舗コード
            ws.cell(row=current_row, column=DeliveryNoteConfig.COL_INDEX_STORE_CODE + 1).value = row_data['store_code']
            
            # 品番
            ws.cell(row=current_row, column=DeliveryNoteConfig.COL_INDEX_PRODUCT_CODE + 1).value = row_data['product_code']
            
            # サイズ
            ws.cell(row=current_row, column=DeliveryNoteConfig.COL_INDEX_SIZE + 1).value = row_data['size']
            
            # カラー
            ws.cell(row=current_row, column=DeliveryNoteConfig.COL_INDEX_COLOR + 1).value = row_data['color']
            
            # 数量
            ws.cell(row=current_row, column=DeliveryNoteConfig.COL_INDEX_QTY + 1).value = row_data['qty']
            
            # 复制样式？
            # 理想情况下应该从上一行(模板行)复制样式
            # 假设模板第8行(index 7)已经有样式
            if idx > 0: # 第一行直接写，第二行开始如果需要复制样式...
                # 暂时简化，只写入数据
                pass

        print(f"Saving output file: {self.output_path}")
        wb.save(self.output_path)