import openpyxl
from openpyxl import load_workbook
import shutil
import os
from config import StoreDetailConfig

class StoreDetailWriter:
    """各店铺明细写入器"""
    
    def __init__(self, template_path, output_path):
        self.template_path = template_path
        self.output_path = output_path
        
    def write(self, data):
        """
        写入数据
        :param data: DataTransformer 的输出结果
        """
        print(f"Writing Store Detail to: {self.output_path}")
        
        # 复制模板
        try:
            shutil.copy(self.template_path, self.output_path)
        except Exception as e:
            print(f"Error copying template: {e}")
            raise e
        
        try:
            wb = load_workbook(self.output_path)
            # 获取模板sheet
            template_sheet = wb.active
            # 假设模板sheet名称为 "Template" 或默认
            # 如果需要保留模板sheet作为复制源，最好不要直接在上面写
            # 但这里我们假设模板sheet就是第一个sheet
            
            # 提取 Brand (从 meta.kanri_no 前3位)
            meta = data.get('metadata', {})
            kanri_no = str(meta.get('kanri_no', '')).strip()
            brand = kanri_no[:3] if len(kanri_no) >= 3 else kanri_no
            
            # 遍历 PT Groups 获取所有店铺数据
            # DataTransformer 返回结构: {'pt_groups': [...], 'skus': [...], ...}
            pt_groups = data.get('pt_groups', [])
            all_skus = data.get('skus', [])
            
            # 收集所有店铺数据并按店铺代码聚合
            # 一个店铺可能在多个PT中出现（虽然理论上一个店铺只有一个配比，但逻辑上可能有多个条目）
            # 这里的需求是“每个店铺一页工作表”，所以我们需要将同一店铺的所有SKU聚合在一起
            # DataTransformer 的输出中，stores 列表里的 store 对象包含了 sku_quantities
            
            stores_map = {} # store_code -> {store_info, items: []}
            
            for group in pt_groups:
                for store in group.get('stores', []):
                    store_code = str(store.get('store_code', ''))
                    
                    if store_code not in stores_map:
                        stores_map[store_code] = {
                            'info': store,
                            'items': []
                        }
                    
                    # 收集该店铺在该PT下的所有有效SKU
                    sku_quantities = store.get('sku_quantities', {})
                    
                    # 使用箱设定A列的逻辑生成Slip No (81 + 4位No)，排除PT前缀
                    # 修正：使用全局顺序号 (G列编号) 生成 Slip No
                    try:
                        # 优先使用 global_seq_no (如果在 data_transformer 中计算了)
                        if 'global_seq_no' in store:
                            seq_no = int(store['global_seq_no'])
                        else:
                            # 降级：尝试使用原始 no (不推荐)
                            seq_no = int(store.get('no', 0))
                        
                        slip_no = f"81{seq_no:04d}"
                    except:
                        slip_no = f"81{str(store.get('no', ''))}"
                    
                    for sku in all_skus:
                        p_code = str(sku.get('product_code', ''))
                        color = str(sku.get('color', ''))
                        size = str(sku.get('size', ''))
                        key = f"{p_code}_{color}_{size}"
                        
                        qty = sku_quantities.get(key, 0)
                        if qty > 0:
                            stores_map[store_code]['items'].append({
                                'slip_no': slip_no,
                                'product_code': p_code,
                                'color': color,
                                'size': size,
                                'qty': qty
                            })

            # 对店铺进行排序
            sorted_store_codes = sorted(stores_map.keys())
            
            # 为每个店铺创建Sheet
            for store_code in sorted_store_codes:
                store_data = stores_map[store_code]
                items = store_data['items']
                if not items:
                    continue
                    
                store_name = store_data['info'].get('store_name', '')
                sheet_title = f"{store_code}_{store_name}"[:30] # Excel sheet name limit 31 chars
                # 移除非法字符
                invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
                for char in invalid_chars:
                    sheet_title = sheet_title.replace(char, '')
                
                # 复制模板Sheet
                target_sheet = wb.copy_worksheet(template_sheet)
                target_sheet.title = sheet_title
                
                # 写入数据
                current_row = StoreDetailConfig.WRITE_START_ROW + 1
                
                # 填充店铺名称 (假设在表头某处，模板分析显示Row 5 Col F是店铺名)
                # [None, '受渡伝票NO', None, None, None, '店舗名', None, None, None] (Row 5 in 0-indexed is Row 6 in Excel? No, analysis says Row 1-5 printed)
                # Analyze output: Row 5: [None, '受渡伝票NO', None, None, None, '店舗名', None, None, None] -> This is index 4.
                # So Row 5 (1-based) is the label row?
                # Actually, let's look at config: WRITE_START_ROW = 7 (index 7, row 8).
                # The analyze output showed:
                # Row 5 (index 4): [..., '店舗名', ...] -> value is '店舗名'
                # Usually we want to write the actual store name next to it or replacing it?
                # Or maybe there is a specific cell for store name.
                # Given I don't have exact coordinates for store name cell, I will stick to writing the list.
                # But user said "一个店铺一页工作表", implies the sheet header should probably have the store name.
                # Let's try to put store name in a reasonable place if template has a placeholder.
                # Based on analysis: Row 5, Col 6 (Index 5) is '店舗名'. Let's write the actual name in Row 5, Col 7 (Index 6) or replace it?
                # Let's assume Row 5 Col 7 (G5) is the place.
                try:
                    target_sheet.cell(row=5, column=7).value = store_name
                except:
                    pass

                # Calculate total quantity for this store
                total_qty = sum(item['qty'] for item in items)
                
                # Write Total Quantity to C4
                target_sheet['C4'] = total_qty
                
                # Write Management No (kanri_no) to C5
                target_sheet['C5'] = kanri_no

                for item in items:
                    # B: Slip No
                    target_sheet.cell(row=current_row, column=StoreDetailConfig.COL_INDEX_SLIP_NO + 1).value = item['slip_no']
                    # C: Brand
                    target_sheet.cell(row=current_row, column=StoreDetailConfig.COL_INDEX_BRAND + 1).value = brand
                    # D: Store Code
                    target_sheet.cell(row=current_row, column=StoreDetailConfig.COL_INDEX_STORE_CODE + 1).value = str(store_code)
                    # E: Product Code
                    target_sheet.cell(row=current_row, column=StoreDetailConfig.COL_INDEX_PRODUCT_CODE + 1).value = item['product_code']
                    # F: Size
                    target_sheet.cell(row=current_row, column=StoreDetailConfig.COL_INDEX_SIZE + 1).value = item['size']
                    # G: Color
                    target_sheet.cell(row=current_row, column=StoreDetailConfig.COL_INDEX_COLOR + 1).value = item['color']
                    # H: Qty
                    target_sheet.cell(row=current_row, column=StoreDetailConfig.COL_INDEX_QTY + 1).value = item['qty']
                    
                    current_row += 1
            
            # 删除原始模板Sheet (如果生成了新Sheet)
            if len(wb.sheetnames) > 1:
                wb.remove(template_sheet)
            
            print(f"Saving Store Detail file: {self.output_path}")
            wb.save(self.output_path)
            return self.output_path
            
        except Exception as e:
            print(f"Error writing store detail: {e}")
            raise e
