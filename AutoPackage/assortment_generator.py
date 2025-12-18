#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
アソート明細生成模块 - 将填写了箱号的配分表转换为アソート明細格式
"""
import openpyxl
from openpyxl import load_workbook
import xlrd
from typing import Dict, List, Optional
import os
from datetime import datetime
from config import TemplateConfig, AssortmentConfig

from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from copy import copy

class AssortmentGenerator:
    """アソート明細生成器"""
    
    def __init__(self, input_path: str, template_path: str, output_path: str, week_num: str = None):
        """
        初始化生成器
        
        Args:
            input_path: 输入文件路径（填写了箱号的配分表）
            template_path: 模板文件路径（アソート明細模板）
            output_path: 输出文件路径
            week_num: 周数 (可选)
        """
        self.input_path = input_path
        self.template_path = template_path
        self.output_path = output_path
        self.data_rows = []
        self.logs = [] # 用于存储日志信息
        self.week_num = week_num

    def process(self):
        """执行转换流程"""
        # 1. 读取输入文件数据
        self._read_input_data()
        
        # 2. 写入到输出模板
        self._write_to_template()
        
        return self.output_path

    def _get_week_number(self, date_val):
        """从日期获取周数"""
        try:
            dt = None
            if isinstance(date_val, datetime):
                dt = date_val
            elif isinstance(date_val, str):
                # 尝试解析常见格式
                for fmt in ['%Y/%m/%d', '%Y-%m-%d', '%Y年%m月%d日']:
                    try:
                        dt = datetime.strptime(date_val, fmt)
                        break
                    except:
                        pass
            
            if dt:
                # ISO 周数
                return f"{dt.isocalendar()[1]:02d}"
        except Exception as e:
            print(f"Error parsing date: {e}")
        
        # 默认返回空或抛出
        return "00"

    def _read_input_data(self):
        """读取输入文件中的数据"""
        print(f"Reading input file: {self.input_path}")
        
        # 根据文件扩展名选择读取方式
        ext = os.path.splitext(self.input_path)[1].lower()
        
        if ext in ['.xlsx', '.xlsm']:
            self._read_xlsx()
        elif ext == '.xls':
            self._read_xls()
        else:
            raise ValueError(f"Unsupported file format: {ext}")
            
    def _read_xlsx(self):
        """使用openpyxl读取xlsx文件"""
        wb = load_workbook(self.input_path, data_only=True)
        
        # 遍历所有PT sheet
        for sheet_name in wb.sheetnames:
            if "PT-" not in sheet_name:
                continue
                
            ws = wb[sheet_name]
            self._process_sheet_data_openpyxl(ws)

    def _process_sheet_data_openpyxl(self, ws):
        """处理单个sheet的数据 (openpyxl)"""
        # 1. 提取元数据
        # E1: Kanri No
        kanri_no = str(ws['E1'].value).strip() if ws['E1'].value else ""
        self.kanri_no = kanri_no # Store for filename
        man_no_prefix = kanri_no[:3] if len(kanri_no) >= 3 else kanri_no
        
        # E4: Store Date -> Week Number
        store_date = ws['E4'].value
        if self.week_num is None and store_date:
             self.week_num = self._get_week_number(store_date)
             self.logs.append(f"Detected Week Number: {self.week_num} from date {store_date}")
        
        current_week = self.week_num or "00"
        
        # 2. 提取SKU信息
        # Row 1 (Index 1): JAN Code (from I1)
        # Row 2 (Index 2): Product Code
        # Row 3 (Index 3): Color
        # Row 4 (Index 4): Size
        
        sku_map = {} # col_index -> {jan, product_code, color, size}
        
        current_col = 8 # Column I
        while True:
            # JAN Code in Row 1
            jan = ws.cell(row=1, column=current_col+1).value
            prod_code = ws.cell(row=2, column=current_col+1).value
            
            if not prod_code: # 如果没有 product code，假设结束
                break
                
            color = ws.cell(row=3, column=current_col+1).value
            size = ws.cell(row=4, column=current_col+1).value
            
            sku_map[current_col] = {
                'jan': str(jan) if jan else "",
                'product_code': str(prod_code),
                'color': str(color) if color is not None else "",
                'size': str(size) if size is not None else ""
            }
            current_col += 1
            
        # 3. 遍历数据行 (从 Row 6 / index 5 开始)
        # Col D (4): Store Code
        # Col E (5): Store Name
        # Col F (6): CTN_NO
        # Col H (8): Total Qty (for validation)
        # Col I (9) ~: SKU Quantities
        
        row_idx = 5 # Row 6
        last_store_code = None
        last_store_name = None
        last_ctn_no = None
        empty_row_count = 0

        while row_idx < ws.max_row:
            row_num = row_idx + 1
            store_code = ws.cell(row=row_num, column=4).value # Col D
            store_name = ws.cell(row=row_num, column=5).value # Col E
            ctn_no_raw = ws.cell(row=row_num, column=6).value # Col F
            total_qty_col = ws.cell(row=row_num, column=8).value # Col H
            
            # Check for totally empty row or "Total" row to break
            first_col_val = ws.cell(row=row_num, column=1).value
            if first_col_val and ("合计" in str(first_col_val) or "Total" in str(first_col_val)):
                break

            # If all key columns are empty, increment empty counter
            has_data = store_code or ctn_no_raw or total_qty_col
            if not has_data and not first_col_val:
                empty_row_count += 1
                if empty_row_count > 10: # Stop after 10 empty rows
                    break
                row_idx += 1
                continue
            
            empty_row_count = 0 # Reset if data found

            # Fill down Store Code and Name
            if store_code:
                last_store_code = store_code
                last_store_name = store_name
            elif last_store_code:
                store_code = last_store_code
                store_name = last_store_name
            
            # Fill down CTN NO (only if store code hasn't changed, logically)
            # But here we just assume if ctn_no is empty, it belongs to the same ctn as previous row
            if ctn_no_raw:
                last_ctn_no = ctn_no_raw
            elif last_ctn_no:
                 # If ctn_no is empty, use the last one
                 ctn_no_raw = last_ctn_no
            
            if not store_code or not ctn_no_raw:
                # If still no store code or ctn no, skip
                row_idx += 1
                continue

            if ctn_no_raw:
                try:
                    ctn_no_str = str(ctn_no_raw).strip()
                    # 尝试转int再补零，或者直接用
                    try:
                        ctn_no_int = int(float(ctn_no_str))
                        ctn_no_formatted = f"{ctn_no_int:04d}"
                    except:
                        ctn_no_formatted = ctn_no_str
                        
                    # 规则: 2位数W + "81" + CTN_NO
                    slip_no = f"{current_week}W81{ctn_no_formatted}"
                    
                    row_sku_sum = 0
                    
                    # 遍历SKU
                    for col_idx, sku_info in sku_map.items():
                        qty = ws.cell(row=row_num, column=col_idx+1).value
                        if qty and isinstance(qty, (int, float)) and qty > 0:
                            qty_int = int(qty)
                            row_sku_sum += qty_int
                            
                            # 构建 Manufacturer Product Code
                            # 管理No前3位 + "-" + 品番 + "-" + 尺码 + "-" + 颜色
                            man_code = f"{man_no_prefix}-{sku_info['product_code']}-{sku_info['size']}-{sku_info['color']}"
                            
                            self.data_rows.append({
                                'delivery_code': str(store_code),
                                'delivery_name': str(store_name),
                                'slip_no': slip_no,
                                'jan': sku_info['jan'],
                                'manufacturer_code': man_code,
                                'qty': qty_int
                            })
                            
                    # 验证合计
                    try:
                        expected_total = int(total_qty_col) if total_qty_col else 0
                        if row_sku_sum != expected_total:
                            msg = f"[Validation Warning] Sheet: {ws.title}, Row: {row_num}, CTN: {ctn_no_formatted} - Sum({row_sku_sum}) != TotalCol({expected_total})"
                            self.logs.append(msg)
                            print(msg)
                    except:
                        pass
                        
                except Exception as e:
                    print(f"Error processing row {row_num}: {e}")
            
            row_idx += 1

    def _read_xls(self):
        """使用xlrd读取xls文件"""
        wb = xlrd.open_workbook(self.input_path)
        
        for sheet in wb.sheets():
            if "PT-" not in sheet.name:
                continue
            self._process_sheet_data_xlrd(sheet)

    def _process_sheet_data_xlrd(self, sheet):
        """处理单个sheet的数据 (xlrd)"""
        # 1. 提取元数据
        try:
            kanri_no = str(sheet.cell_value(0, 4)).strip() # E1
            self.kanri_no = kanri_no # Store for filename
            man_no_prefix = kanri_no[:3] if len(kanri_no) >= 3 else kanri_no
            
            # E4 date
            store_date = sheet.cell_value(3, 4) # Row 4 (index 3), Col E (index 4)
            if self.week_num is None and store_date:
                # xlrd date is float
                if isinstance(store_date, float):
                    dt = xlrd.xldate_as_datetime(store_date, 0)
                    self.week_num = f"{dt.isocalendar()[1]:02d}"
                    self.logs.append(f"Detected Week Number: {self.week_num}")
                else:
                    self.week_num = self._get_week_number(store_date)
        except:
            man_no_prefix = ""
            
        current_week = self.week_num or "00"
        
        # 2. 提取SKU信息
        sku_map = {}
        current_col = 8 # Col I
        
        while current_col < sheet.ncols:
            try:
                # Row 1: JAN
                jan = sheet.cell_value(0, current_col)
                prod_code = sheet.cell_value(1, current_col) # Row 2
                if not prod_code:
                    break
                
                color = sheet.cell_value(2, current_col) # Row 3
                size = sheet.cell_value(3, current_col) # Row 4
                
                # Format
                if isinstance(color, float): color = str(int(color))
                if isinstance(size, float): size = str(int(size))
                if isinstance(prod_code, float): prod_code = str(int(prod_code))
                if isinstance(jan, float): jan = str(int(jan))
                
                sku_map[current_col] = {
                    'jan': str(jan) if jan else "",
                    'product_code': str(prod_code),
                    'color': str(color),
                    'size': str(size)
                }
                current_col += 1
            except IndexError:
                break
                
        # 3. 遍历数据
        row_idx = 5 # Row 6
        last_store_code = None
        last_store_name = None
        last_ctn_no = None
        empty_row_count = 0

        while row_idx < sheet.nrows:
            try:
                store_code = sheet.cell_value(row_idx, 3) # Col D
                store_name = sheet.cell_value(row_idx, 4) # Col E
                ctn_no_raw = sheet.cell_value(row_idx, 5) # Col F
                total_qty_col = sheet.cell_value(row_idx, 7) # Col H
                
                # Check end
                first_val = sheet.cell_value(row_idx, 0)
                if "合计" in str(first_val) or "Total" in str(first_val):
                    break

                # Check empty
                has_data = store_code or ctn_no_raw or total_qty_col
                if not has_data and not first_val:
                    empty_row_count += 1
                    if empty_row_count > 10: break
                    row_idx += 1
                    continue
                empty_row_count = 0

                # Fill down
                if store_code:
                    last_store_code = store_code
                    last_store_name = store_name
                elif last_store_code:
                    store_code = last_store_code
                    store_name = last_store_name
                
                if ctn_no_raw:
                    last_ctn_no = ctn_no_raw
                elif last_ctn_no:
                    ctn_no_raw = last_ctn_no
                
                if not store_code or not ctn_no_raw:
                    row_idx += 1
                    continue
                     
                if ctn_no_raw:
                    try:
                        # CTN NO
                        if isinstance(ctn_no_raw, float):
                            ctn_no_int = int(ctn_no_raw)
                            ctn_no_formatted = f"{ctn_no_int:04d}"
                        else:
                            ctn_no_formatted = str(ctn_no_raw).strip()
                            
                        slip_no = f"{current_week}W81{ctn_no_formatted}"
                        
                        row_sku_sum = 0
                        
                        for col_idx, sku_info in sku_map.items():
                            qty = sheet.cell_value(row_idx, col_idx)
                            if qty and isinstance(qty, (int, float)) and qty > 0:
                                qty_int = int(qty)
                                row_sku_sum += qty_int
                                
                                man_code = f"{man_no_prefix}-{sku_info['product_code']}-{sku_info['size']}-{sku_info['color']}"
                                
                                self.data_rows.append({
                                    'delivery_code': str(store_code) if not isinstance(store_code, float) else str(int(store_code)),
                                    'delivery_name': str(store_name),
                                    'slip_no': slip_no,
                                    'jan': sku_info['jan'],
                                    'manufacturer_code': man_code,
                                    'qty': qty_int
                                })
                        
                        # Validation
                        try:
                            expected_total = int(total_qty_col) if total_qty_col else 0
                            if row_sku_sum != expected_total:
                                msg = f"[Validation Warning] Sheet: PT, Row: {row_idx+1}, CTN: {ctn_no_formatted} - Sum({row_sku_sum}) != TotalCol({expected_total})"
                                self.logs.append(msg)
                        except: pass
                        
                    except ValueError:
                        pass
            except IndexError:
                break
            
            row_idx += 1

    def _write_to_template(self):
        """写入数据到输出模板"""
        print(f"Writing to template: {self.template_path}")
        
        # Check xls
        if self.template_path.lower().endswith('.xls'):
            xlsx_path = self.template_path + "x"
            if os.path.exists(xlsx_path):
                self.template_path = xlsx_path
            else:
                 raise ValueError("Template must be .xlsx format. Please provide .xlsx template.")

        wb = load_workbook(self.template_path)
        ws = wb.active 
        
        start_row = AssortmentConfig.WRITE_START_ROW + 1 
        
        for idx, row_data in enumerate(self.data_rows):
            current_row = start_row + idx
            
            # 届け先コード
            ws.cell(row=current_row, column=AssortmentConfig.COL_INDEX_DELIVERY_CODE + 1).value = row_data['delivery_code']
            
            # 届け先名
            ws.cell(row=current_row, column=AssortmentConfig.COL_INDEX_DELIVERY_NAME + 1).value = row_data['delivery_name']
            
            # 受渡伝票
            ws.cell(row=current_row, column=AssortmentConfig.COL_INDEX_SLIP_NO + 1).value = row_data['slip_no']
            
            # JANコード
            ws.cell(row=current_row, column=AssortmentConfig.COL_INDEX_JAN + 1).value = row_data['jan']
            
            # メーカー品番
            ws.cell(row=current_row, column=AssortmentConfig.COL_INDEX_MANUFACTURER_CODE + 1).value = row_data['manufacturer_code']
            
            # 汇总 (数量)
            ws.cell(row=current_row, column=AssortmentConfig.COL_INDEX_QTY + 1).value = row_data['qty']
            
        # 添加合计行
        if self.data_rows:
            total_row = start_row + len(self.data_rows)
            # Copy style from header row (row 2)
            header_row = 2
            
            # 合计Label (E列 / COL_INDEX_JAN)
            cell_label = ws.cell(row=total_row, column=AssortmentConfig.COL_INDEX_JAN + 1)
            cell_label.value = "合計"
            self._copy_style(ws.cell(row=header_row, column=AssortmentConfig.COL_INDEX_JAN + 1), cell_label)
            
            # 合计数量 (G列 / COL_INDEX_QTY)
            total_qty = sum(r['qty'] for r in self.data_rows)
            cell_qty = ws.cell(row=total_row, column=AssortmentConfig.COL_INDEX_QTY + 1)
            cell_qty.value = total_qty
            self._copy_style(ws.cell(row=header_row, column=AssortmentConfig.COL_INDEX_QTY + 1), cell_qty)
            
            # Apply border to other cells in total row for consistency? 
            # User asked for "header style", usually implies background color and bold text.
            # Let's apply to the whole row range B-G
            for col_idx in range(AssortmentConfig.COL_INDEX_DELIVERY_CODE, AssortmentConfig.COL_INDEX_QTY + 1):
                cell = ws.cell(row=total_row, column=col_idx + 1)
                self._copy_style(ws.cell(row=header_row, column=col_idx + 1), cell)
                if col_idx == AssortmentConfig.COL_INDEX_JAN:
                     cell.value = "合計"
                elif col_idx == AssortmentConfig.COL_INDEX_QTY:
                     cell.value = total_qty
                else:
                     cell.value = None # Clear other cells

        print(f"Saving output file: {self.output_path}")
        wb.save(self.output_path)

    def _copy_style(self, source_cell, target_cell):
        """复制单元格样式"""
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)