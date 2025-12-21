#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
模板写入模块 - 基于openpyxl实现，支持合并单元格
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
from config import TemplateConfig
from typing import Dict, List


class TemplateWriter:
    """基于openpyxl的Excel文件写入器"""
    
    def __init__(self, template_path: str, output_path: str):
        """
        初始化写入器
        
        Args:
            template_path: 模板文件路径
            output_path: 输出文件路径
        """
        self.template_path = template_path
        self.output_path = output_path
        self.workbook = None
        
    def write(self, transformed_data: Dict) -> str:
        """
        写入数据到输出文件
        
        Args:
            transformed_data: 转换后的数据
            
        Returns:
            输出文件路径
        """
        try:
            # 加载模板
            print(f"加载模板: {self.template_path}")
            self.workbook = load_workbook(self.template_path)
            
            # 重命名模板PT-1页以避免名称冲突
            template_sheet_name = TemplateConfig.PT_TEMPLATE_SHEET
            temp_template_name = "_TEMPLATE_PT_TEMP"
            template_pt_sheet = None
            
            if template_sheet_name in self.workbook.sheetnames:
                template_pt_sheet = self.workbook[template_sheet_name]
                template_pt_sheet.title = temp_template_name
            
            # 更新商品一覧页
            self._update_product_list(transformed_data)
            
            # 创建所有PT页
            self._create_pt_sheets(transformed_data, template_pt_sheet)
            
            # 删除临时模板页
            if template_pt_sheet and temp_template_name in self.workbook.sheetnames:
                self.workbook.remove(template_pt_sheet)
            
            # 保存文件
            print(f"保存输出文件: {self.output_path}")
            self.workbook.save(self.output_path)
            
            return self.output_path
            
        except Exception as e:
            raise Exception(f"写入文件失败: {e}")
    
    def _update_product_list(self, data: Dict):
        """更新商品一覧页"""
        try:
            if TemplateConfig.PRODUCT_LIST_SHEET not in self.workbook.sheetnames:
                raise Exception("找不到商品一覧sheet")
            
            sheet = self.workbook[TemplateConfig.PRODUCT_LIST_SHEET]
            
            # 更新管理No
            kanri_no = data.get('metadata', {}).get('kanri_no', '')
            if kanri_no:
                kanri_cell = sheet.cell(
                    row=TemplateConfig.PRODUCT_LIST_KANRI_NO_ROW + 1,
                    column=TemplateConfig.PRODUCT_LIST_KANRI_NO_COL + 1
                )
                kanri_cell.value = kanri_no
            
            # 写入SKU数据（保持模板样式）
            skus = data['skus']
            # 获取模板行样式
            template_row = TemplateConfig.PRODUCT_LIST_DATA_START + 1
            
            # 统计总出荷数
            grand_total_ship_qty = 0
            
            # 确定列范围（从No列到增产数列）
            # 注意：这里的列索引是0-indexed，所以需要+1
            start_col = TemplateConfig.PRODUCT_LIST_COL_NO
            end_col = getattr(TemplateConfig, 'PRODUCT_LIST_COL_INCREASE', 8) # 默认为8
            
            last_row_num = 0
            
            for idx, sku in enumerate(skus):
                row_num = TemplateConfig.PRODUCT_LIST_DATA_START + 1 + idx
                last_row_num = row_num
                
                # 遍历所有列，确保边框完整
                for col_idx in range(start_col, end_col + 1):
                    cell = sheet.cell(row=row_num, column=col_idx + 1)
                    
                    # 根据列索引写入数据
                    val = None
                    if col_idx == TemplateConfig.PRODUCT_LIST_COL_NO:
                        val = idx + 1
                    elif col_idx == TemplateConfig.PRODUCT_LIST_COL_CODE:
                        val = str(sku['product_code'])
                    elif col_idx == TemplateConfig.PRODUCT_LIST_COL_COLOR:
                        val = str(sku['color'])
                    elif col_idx == TemplateConfig.PRODUCT_LIST_COL_SIZE:
                        val = str(sku['size'])
                    elif col_idx == TemplateConfig.PRODUCT_LIST_COL_JAN:
                        val = str(sku.get('jan_code', ''))
                    elif col_idx == getattr(TemplateConfig, 'PRODUCT_LIST_COL_SHIP_QTY', 6):
                        qty = sku.get('total_qty', 0)
                        if qty:
                            val = int(qty)
                            grand_total_ship_qty += val
                    
                    if val is not None:
                        cell.value = val
                    
                    # 复制样式（所有列都应用样式，从而确保边框）
                    self._copy_cell_style(sheet, template_row, col_idx + 1, cell)
            
            # 写入合计行
            total_row_num = last_row_num + 1
            
            # 定义合计行样式（灰色背景，加粗，边框）
            total_style = {
                'font': Font(name='ＭＳ Ｐゴシック', size=10, bold=True), # 稍微大一点
                'alignment': Alignment(horizontal='center', vertical='center'),
                'fill': PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid'),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            }
            
            # 合计行从No列到增产数列
            for col_idx in range(start_col, end_col + 1):
                cell = sheet.cell(row=total_row_num, column=col_idx + 1)
                
                # JAN列写"合計"
                if col_idx == TemplateConfig.PRODUCT_LIST_COL_JAN:
                    cell.value = "合計"
                # 出荷数列写总数
                elif col_idx == getattr(TemplateConfig, 'PRODUCT_LIST_COL_SHIP_QTY', 6):
                    cell.value = grand_total_ship_qty
                else:
                    cell.value = "" # 其他列留空，但应用样式
                
                # 应用样式
                cell.font = total_style['font']
                cell.alignment = total_style['alignment']
                cell.fill = total_style['fill']
                cell.border = total_style['border']
                
        except Exception as e:
            print(f"警告: 更新商品一覧页失败: {e}")
    
    def _copy_cell_style(self, sheet, template_row, template_col, target_cell, override_font_name=None, override_font_size=None):
        """从模板单元格复制样式到目标单元格"""
        try:
            template_cell = sheet.cell(row=template_row, column=template_col)
            if template_cell.font:
                new_font = copy(template_cell.font)
                if override_font_name:
                    new_font.name = override_font_name
                if override_font_size:
                    new_font.size = override_font_size
                target_cell.font = new_font
            if template_cell.border:
                target_cell.border = copy(template_cell.border)
            if template_cell.fill:
                target_cell.fill = copy(template_cell.fill)
            if template_cell.alignment:
                target_cell.alignment = copy(template_cell.alignment)
        except:
            pass
    
    def _create_pt_sheets(self, data: Dict, template_sheet):
        """创建所有PT页"""
        pt_groups = data['pt_groups']
        metadata = data.get('metadata', {})
        skus = data['skus']
        
        for pt_group in pt_groups:
            self._create_single_pt_sheet(pt_group, metadata, skus, template_sheet)
    
    def _create_single_pt_sheet(self, pt_group: Dict, metadata: Dict, skus: List[Dict], template_sheet):
        """创建单个PT页"""
        try:
            pt_name = pt_group['pt_name']
            print(f"创建 {pt_name}...")
            
            # 复制模板sheet
            if template_sheet:
                new_sheet = self.workbook.copy_worksheet(template_sheet)
                new_sheet.title = pt_name
            else:
                new_sheet = self.workbook.create_sheet(pt_name)
            
            # 设置列宽
            # A, B, C, D (1-4) 列宽为 5
            for col_idx in range(1, 5):
                col_letter = get_column_letter(col_idx)
                new_sheet.column_dimensions[col_letter].width = 5
            
            # E (5) 店铺名列宽为 27
            new_sheet.column_dimensions['E'].width = 27
            
            # 写入表头
            self._write_pt_header(new_sheet, pt_group, metadata, skus)
            
            # 写入数据
            self._write_pt_data(new_sheet, pt_group, skus)
            
        except Exception as e:
            print(f"警告: 创建PT页 {pt_group.get('pt_name', '?')} 失败: {e}")
            import traceback
            traceback.print_exc()
    
    def _write_pt_header(self, sheet, pt_group: Dict, metadata: Dict, skus: List[Dict]):
        """写入PT页表头并实现合并单元格"""
        pt_name = pt_group['pt_name']
        total_qty = pt_group['total_qty']
        kanri_no = metadata.get('kanri_no', '')
        store_date = metadata.get('store_date', metadata.get('delivery_date', ''))
        
        # 定义样式（用于超出模板列的情况）
        # 表头SKU列样式（灰色背景）
        header_sku_style = {
            'font': Font(name='ＭＳ Ｐゴシック', size=9),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'fill': PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        
        # 统一字体
        common_font = Font(name='ＭＳ Ｐゴシック', size=9)
        
        # 行1 (索引0): 管理No区域 - 合并A1:D1
        sheet.merge_cells('A1:D1')
        sheet['A1'].value = "管理No"
        sheet['A1'].font = common_font
        sheet['E1'].value = kanri_no
        sheet['E1'].font = common_font
        sheet['E1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 行1: No.标签 (F1)
        sheet['F1'].value = "Jan"
        sheet['F1'].font = common_font
        
        # 行1: SKU序号 (从I列开始)
        for idx, sku in enumerate(skus):
            col_letter = get_column_letter(TemplateConfig.PT_SKU_START_COL + 1 + idx)
            cell = sheet[f'{col_letter}1']
            cell.value = str(sku.get('jan_code', ''))
            # 应用样式
            cell.font = header_sku_style['font']
            # JANCODE 需要换行
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = header_sku_style['fill']
            cell.border = header_sku_style['border']
        
        # 行2: パターン区域 - 合并A2:D2
        sheet.merge_cells('A2:D2')
        sheet['A2'].value = "パターン"
        sheet['A2'].font = common_font
        sheet['E2'].value = pt_name
        sheet['E2'].font = common_font
        sheet['E2'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 行2: 品番标签
        sheet['F2'].value = "品番"
        sheet['F2'].font = common_font
        
        # 行2: 品番序列
        for idx, sku in enumerate(skus):
            col_letter = get_column_letter(TemplateConfig.PT_SKU_START_COL + 1 + idx)
            cell = sheet[f'{col_letter}2']
            cell.value = str(sku['product_code'])
            # 应用样式
            cell.font = header_sku_style['font']
            cell.alignment = header_sku_style['alignment']
            cell.fill = header_sku_style['fill']
            cell.border = header_sku_style['border']
        
        # 行3: 枚数区域 - 合并A3:D3
        sheet.merge_cells('A3:D3')
        sheet['A3'].value = "枚数"
        sheet['A3'].font = common_font
        sheet['E3'].value = total_qty
        sheet['E3'].font = common_font
        sheet['E3'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 行3: カラー标签
        sheet['F3'].value = "カラー"
        sheet['F3'].font = common_font
        
        # 行3: カラー序列
        for idx, sku in enumerate(skus):
            col_letter = get_column_letter(TemplateConfig.PT_SKU_START_COL + 1 + idx)
            cell = sheet[f'{col_letter}3']
            cell.value = str(sku['color'])
            # 应用样式
            cell.font = header_sku_style['font']
            cell.alignment = header_sku_style['alignment']
            cell.fill = header_sku_style['fill']
            cell.border = header_sku_style['border']
        
        # 行4: 納期区域 - 合并A4:D4
        sheet.merge_cells('A4:D4')
        sheet['A4'].value = "納期"
        sheet['A4'].font = common_font
        sheet['E4'].value = store_date
        sheet['E4'].font = common_font
        sheet['E4'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 行4: サイズ标签
        sheet['F4'].value = "サイズ"
        sheet['F4'].font = common_font
        
        # 行4: サイズ序列
        for idx, sku in enumerate(skus):
            col_letter = get_column_letter(TemplateConfig.PT_SKU_START_COL + 1 + idx)
            cell = sheet[f'{col_letter}4']
            cell.value = str(sku['size'])
            # 应用样式
            cell.font = header_sku_style['font']
            cell.alignment = header_sku_style['alignment']
            cell.fill = header_sku_style['fill']
            cell.border = header_sku_style['border']
        
        # 行5: 数据列表头
        headers = ["No.", "タイプ", "ランク", "コード", "店舗名", "CTN_NO", "パターン", "合計"]
        for col_idx, header in enumerate(headers):
            cell = sheet.cell(row=5, column=col_idx + 1)
            cell.value = header
            # 这里也应该统一字体，但通常表头有特定样式，暂时应用common_font或保持原样
            # 假设表头样式由模板决定，这里只写入值。
            # 但用户要求"写入的内容的字体统一"，所以最好也设置一下，或者依赖模板但修改字体名
            # 由于没有_copy_cell_style，这里直接设置
            cell.font = Font(name='ＭＳ Ｐゴシック', size=9) # 假设表头稍微小一点或者和正文一样
        
        # 行5: SKU列的列标题（也需要灰色背景）
        for idx in range(len(skus)):
            col_num = TemplateConfig.PT_SKU_START_COL + 1 + idx
            col_letter = get_column_letter(col_num)
            cell = sheet[f'{col_letter}5']
            # 可以留空或写编号
            cell.value = ""
            # 应用灰色背景样式
            cell.font = header_sku_style['font']
            cell.alignment = header_sku_style['alignment']
            cell.fill = header_sku_style['fill']
            cell.border = header_sku_style['border']

    
    def _write_pt_data(self, sheet, pt_group: Dict, skus: List[Dict]):
        """写入PT页数据"""
        stores = pt_group['stores']
        template_row = 6  # 使用第6行作为样式模板
        pt_name = pt_group['pt_name']
        
        # 定义SKU数据列样式（白色背景，只有边框）
        data_sku_style = {
            'font': Font(name='ＭＳ Ｐゴシック', size=9),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'fill': PatternFill(fill_type=None),  # 无背景色（白色）
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        
        # 统计总数
        grand_total_qty = 0
        sku_totals = {idx: 0 for idx in range(len(skus))}
        last_row_num = TemplateConfig.PT_DATA_START_ROW
        
        for idx, store in enumerate(stores):
            row_num = TemplateConfig.PT_DATA_START_ROW + 1 + idx
            last_row_num = row_num
            
            # No. - Generate 4 digit ID (PT_NUM + 3 digit Local Index)
            try:
                # Extract PT number (e.g., PT-1 -> 1)
                pt_num = int(pt_name.split('-')[1])
                # Use PT local index instead of original No to ensure sequence
                local_idx = store.get('pt_local_idx', idx + 1)
                new_id = f"{pt_num}{local_idx:03d}"
            except:
                new_id = store.get('no', '')
            
            cell = sheet.cell(row=row_num, column=1)
            cell.value = new_id
            self._copy_cell_style(sheet, template_row, 1, cell, override_font_name='ＭＳ Ｐゴシック', override_font_size=9)
            
            # タイプ
            cell = sheet.cell(row=row_num, column=2)
            cell.value = store['type']
            self._copy_cell_style(sheet, template_row, 2, cell, override_font_name='ＭＳ Ｐゴシック', override_font_size=9)
            
            # ランク
            cell = sheet.cell(row=row_num, column=3)
            cell.value = store.get('rank', '')
            self._copy_cell_style(sheet, template_row, 3, cell, override_font_name='ＭＳ Ｐゴシック', override_font_size=9)
            
            # コード
            cell = sheet.cell(row=row_num, column=4)
            cell.value = store['store_code']
            self._copy_cell_style(sheet, template_row, 4, cell, override_font_name='ＭＳ Ｐゴシック', override_font_size=9)
            
            # 店舗名
            cell = sheet.cell(row=row_num, column=5)
            cell.value = store['store_name']
            self._copy_cell_style(sheet, template_row, 5, cell, override_font_name='ＭＳ Ｐゴシック', override_font_size=9)
            
            # CTN_NO（黄色背景）- 用户要求在G列填入自动生成的序号(0001顺序)
            cell = sheet.cell(row=row_num, column=6)
            # cell.value = store['ctn_no'] # 暂时注释掉
            cell.value = "" # F列 (CTN_NO标题) 留空
            self._copy_cell_style(sheet, template_row, 6, cell, override_font_name='ＭＳ Ｐゴシック', override_font_size=9)
            
            # パターン (G列) -> 现在用来填入 全局顺序号
            cell = sheet.cell(row=row_num, column=7)
            global_seq = store.get('global_seq_no', 0)
            cell.value = f"{global_seq:04d}" if global_seq else ""
            self._copy_cell_style(sheet, template_row, 7, cell, override_font_name='ＭＳ Ｐゴシック', override_font_size=9)
            # 确保G列也是居中对齐
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 合計
            store_total = store['total_qty']
            grand_total_qty += store_total
            cell = sheet.cell(row=row_num, column=8)
            cell.value = store_total
            self._copy_cell_style(sheet, template_row, 8, cell, override_font_name='ＭＳ Ｐゴシック', override_font_size=9)
            
            # SKU数量 - 白色背景，仅边框
            for sku_idx, sku in enumerate(skus):
                col_num = TemplateConfig.PT_COL_FIRST_SKU + 1 + sku_idx
                sku_key = f"{sku['product_code']}_{sku['color']}_{sku['size']}"
                qty = store['sku_quantities'].get(sku_key, 0)
                sku_totals[sku_idx] += qty
                
                cell = sheet.cell(row=row_num, column=col_num)
                if qty > 0:
                    cell.value = int(qty)
                else:
                    cell.value = ""
                
                # 应用样式：白色背景+ 边框
                cell.font = data_sku_style['font']
                cell.alignment = data_sku_style['alignment']
                cell.fill = data_sku_style['fill']
                cell.border = data_sku_style['border']
        
        # 写入最后一行合计
        total_row_num = last_row_num + 1
        
        # 定义合计行样式（灰色背景，居中，边框）
        total_style = {
            'font': Font(name='ＭＳ Ｐゴシック', size=9, bold=True),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'fill': PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        
        # 写入总合计
        cell = sheet.cell(row=total_row_num, column=TemplateConfig.PT_COL_TOTAL + 1)
        cell.value = grand_total_qty
        cell.font = total_style['font']
        cell.alignment = total_style['alignment']
        cell.fill = total_style['fill']
        cell.border = total_style['border']
