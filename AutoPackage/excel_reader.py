#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel读取模块 - 读取配分表数据
"""
import xlrd
import openpyxl
from config import AllocationTableConfig, DetailTableConfig, TemplateConfig
from typing import Dict, List, Tuple
import pandas as pd
from openpyxl import load_workbook
import logging

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class BoxSettingReader:
    """读取工厂返回的箱设定明细表（经过人工分箱处理）"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        
    def read(self) -> List[Dict]:
        """
        读取所有PT页的箱明细数据
        
        Returns:
            List[Dict]: 箱数据列表
        """
        try:
            logger.info(f"Loading box setting file: {self.file_path}")
            wb = load_workbook(self.file_path, data_only=True)
            all_boxes = []
            
            for sheet_name in wb.sheetnames:
                # 跳过商品一覧页
                if sheet_name == TemplateConfig.PRODUCT_LIST_SHEET:
                    continue
                
                sheet = wb[sheet_name]
                
                # 验证是否为PT页 (检查表头 A5)
                # 1-indexed: Row 5, Col 1
                header_cell = sheet.cell(row=TemplateConfig.PT_DATA_HEADER_ROW + 1, column=TemplateConfig.PT_COL_NO + 1)
                # 注意：TemplateWriter可能写入的是"No."，也可能只是该行开始数据
                # Check header value explicitly
                # Writer: sheet.cell(row=5, column=1).value = "No."
                if header_cell.value != "No.":
                    logger.debug(f"Skipping sheet {sheet_name}: Not a PT sheet (A5 != No.)")
                    continue
                    
                logger.info(f"Processing sheet: {sheet_name}")
                sheet_boxes = self._process_sheet(sheet, sheet_name)
                all_boxes.extend(sheet_boxes)
            
            logger.info(f"Total boxes loaded: {len(all_boxes)}")
            return all_boxes
            
        except Exception as e:
            logger.error(f"Error reading box setting file: {e}")
            raise

    def _process_sheet(self, sheet, sheet_name) -> List[Dict]:
        """处理单个PT页"""
        # 1. 读取SKU列定义
        skus = []
        col_idx = 0
        while True:
            # 1-indexed column
            col = TemplateConfig.PT_SKU_START_COL + 1 + col_idx
            
            # 检查是否有品番 (Row 2)
            prod_code = sheet.cell(row=2, column=col).value
            if not prod_code:
                break
                
            skus.append({
                'product_code': str(prod_code),
                'jan_code': str(sheet.cell(row=1, column=col).value or ''),
                'color': str(sheet.cell(row=3, column=col).value or ''),
                'size': str(sheet.cell(row=4, column=col).value or '')
            })
            col_idx += 1
            
        # 2. 读取元数据
        # E1: 管理No
        kanri_no = str(sheet.cell(row=1, column=5).value or '')
        # E4: 納期 (这里作为 Store Date / Delivery Date)
        delivery_date = str(sheet.cell(row=4, column=5).value or '')
        
        # 3. 读取数据行
        box_map = {} # Key: ctn_no -> BoxData
        
        row = TemplateConfig.PT_DATA_START_ROW + 1 # Row 6
        last_ctn_no = None
        last_store_code = None
        last_store_name = None
        last_pattern = None
        
        while row <= sheet.max_row:
            # 检查 CTN_NO 列 (F列, Col 6)
            ctn_no_val = sheet.cell(row=row, column=TemplateConfig.PT_COL_CTN_NO + 1).value
            store_code_val = sheet.cell(row=row, column=TemplateConfig.PT_COL_STORE_CODE + 1).value
            
            # 如果整行关键数据为空，可能是空行，也可能是结束
            if ctn_no_val is None and store_code_val is None:
                # 检查是否所有SKU数据也为空，如果是，则跳过
                is_empty = True
                for i in range(len(skus)):
                    if sheet.cell(row=row, column=TemplateConfig.PT_SKU_START_COL + 1 + i).value:
                        is_empty = False
                        break
                
                if is_empty:
                    # 连续空行超过10行则停止
                    # 这里简单判断：如果后面还有数据则继续，否则结束？
                    # 为简单起见，如果 A列 No. 也没了，就结束
                    if sheet.cell(row=row, column=1).value is None:
                        # 再次确认下一行，防止中间空行
                        if sheet.cell(row=row+1, column=1).value is None:
                            break
                    row += 1
                    continue
                
                # 如果有SKU数据但没有CTN/Store，假设是上一行的延续
                if last_ctn_no is not None:
                    ctn_no_val = last_ctn_no
                    store_code_val = last_store_code
                else:
                    row += 1
                    continue
            else:
                # 更新上下文
                last_ctn_no = ctn_no_val
                last_store_code = store_code_val
                last_store_name = sheet.cell(row=row, column=TemplateConfig.PT_COL_STORE_NAME + 1).value
                last_pattern = sheet.cell(row=row, column=TemplateConfig.PT_COL_PATTERN + 1).value
            
            # 确保有 ctn_no
            if not last_ctn_no:
                row += 1
                continue
                
            key = str(last_ctn_no)
            
            if key not in box_map:
                dept = kanri_no[:3] if len(kanri_no) >= 3 else kanri_no
                
                box_map[key] = {
                    'ctn_no': last_ctn_no,
                    'store_code': last_store_code,
                    'store_name': last_store_name,
                    'pattern': last_pattern,
                    'delivery_date': delivery_date, # 出区日
                    'store_date': delivery_date,    # 店着日 (暂同)
                    'dept': dept,
                    'kanri_no': kanri_no,
                    'total_qty': 0,
                    'items': []
                }
            
            # 读取SKU数据
            for i, sku in enumerate(skus):
                qty_col = TemplateConfig.PT_SKU_START_COL + 1 + i
                qty_val = sheet.cell(row=row, column=qty_col).value
                
                if qty_val and isinstance(qty_val, (int, float)) and qty_val > 0:
                    qty = int(qty_val)
                    # 查找是否已存在该SKU (合并多行情况)
                    existing_item = next((item for item in box_map[key]['items'] 
                                        if item['maker_code'] == f"{sku['product_code']}-{sku['color']}-{sku['size']}"), None)
                    
                    if existing_item:
                        existing_item['qty'] += qty
                    else:
                        box_map[key]['items'].append({
                            'maker_code': f"{sku['product_code']}-{sku['color']}-{sku['size']}",
                            'product_name': '',
                            'qty': qty
                        })
                    
                    box_map[key]['total_qty'] += qty
            
            row += 1
            
        return list(box_map.values())


class DetailTableReader:
    """明细表读取器"""
    
    @staticmethod
    def read_jan_map(file_path: str) -> Dict[Tuple[str, str, str], str]:
        """
        读取明细表并返回 JANCODE 映射字典
        
        Args:
            file_path: 明细表文件路径
            
        Returns:
            Dict: {(品番, カラー, サイズ): JANCODE}
        """
        try:
            # 智能判断文件类型读取
            file_path_lower = file_path.lower()
            
            if file_path_lower.endswith('.csv') or file_path_lower.endswith('.txt'):
                # 尝试不同的分隔符和编码
                try:
                    # 优先尝试制表符分隔（常见于系统导出）
                    df = pd.read_csv(file_path, sep='\t', dtype=str)
                    # 如果只有一列，可能不是制表符，尝试逗号
                    if len(df.columns) < 2:
                        df = pd.read_csv(file_path, sep=',', dtype=str)
                except:
                    # 如果UTF-8失败，尝试GBK/Shift-JIS
                    try:
                        df = pd.read_csv(file_path, sep='\t', encoding='shift-jis', dtype=str)
                    except:
                        df = pd.read_csv(file_path, sep='\t', encoding='gbk', dtype=str)
            else:
                # 默认为 Excel
                df = pd.read_excel(file_path, dtype=str)
            
            # 清理列名（去除空白）
            df.columns = df.columns.str.strip()
            
            # 检查必要的列是否存在
            required_cols = [
                DetailTableConfig.COL_PRODUCT_CODE,
                DetailTableConfig.COL_COLOR,
                DetailTableConfig.COL_SIZE,
                DetailTableConfig.COL_JAN
            ]
            
            for col in required_cols:
                if col not in df.columns:
                    raise ValueError(f"明细表中缺少列: {col}")
            
            # 构建映射字典
            jan_map = {}
            for _, row in df.iterrows():
                # 转换为字符串并去除空白
                product_code = str(row[DetailTableConfig.COL_PRODUCT_CODE]).strip()
                # 处理颜色：去除小数点（如果是浮点数）
                color_val = row[DetailTableConfig.COL_COLOR]
                # 由于读取时指定了 dtype=str，这里通常是字符串，但为了保险起见保留处理逻辑
                # 如果是字符串形式的浮点数 "403.0"，split处理
                if isinstance(color_val, str):
                    if '.' in color_val:
                        try:
                            color = str(int(float(color_val)))
                        except:
                            color = color_val.strip()
                    else:
                        color = color_val.strip()
                elif isinstance(color_val, float) and color_val.is_integer():
                    color = str(int(color_val))
                else:
                    color = str(color_val).strip()
                    
                # 处理尺码
                size_val = row[DetailTableConfig.COL_SIZE]
                if isinstance(size_val, str):
                    if '.' in size_val:
                         try:
                            size = str(int(float(size_val)))
                         except:
                            size = size_val.strip()
                    else:
                        size = size_val.strip()
                elif isinstance(size_val, float) and size_val.is_integer():
                    size = str(int(size_val))
                else:
                    size = str(size_val).strip()
                
                # 处理JAN：去除小数点
                jan_val = row[DetailTableConfig.COL_JAN]
                if pd.notna(jan_val):
                    if isinstance(jan_val, str):
                        if '.' in jan_val:
                            try:
                                jan = str(int(float(jan_val)))
                            except:
                                jan = jan_val.strip()
                        else:
                            jan = jan_val.strip()
                    elif isinstance(jan_val, float):
                        jan = str(int(jan_val))
                    else:
                        jan = str(jan_val).strip()
                else:
                    jan = ""
                
                key = (product_code, color, size)
                if jan:
                    jan_map[key] = jan
            
            return jan_map
            
        except Exception as e:
            raise Exception(f"读取明细表失败: {e}")


class AllocationTableReader:
    """配分表读取器"""
    
    def __init__(self, file_path: str):
        """
        初始化读取器
        
        Args:
            file_path: 配分表文件路径
        """
        self.file_path = file_path
        self.workbook = None
        self.metadata = {}
        self.products_data = []  # 存储所有品番的数据
        self.engine = 'xlrd' # 'xlrd' or 'openpyxl'
        
    def read(self) -> Dict:
        """
        读取配分表文件
        
        Returns:
            包含所有数据的字典
        """
        try:
            if self.file_path.lower().endswith('.xlsx'):
                return self._read_xlsx()
            else:
                return self._read_xls()
                
        except Exception as e:
            raise Exception(f"读取配分表文件失败: {e}")

    def _read_xls(self) -> Dict:
        """使用xlrd读取.xls文件"""
        self.engine = 'xlrd'
        self.workbook = xlrd.open_workbook(self.file_path, formatting_info=False)
        
        # 读取所有品番sheet（跳过可能的汇总sheet）
        for sheet_idx in range(self.workbook.nsheets):
            sheet = self.workbook.sheet_by_index(sheet_idx)
            sheet_name = sheet.name
            
            # 跳过空sheet或太小的sheet
            if sheet.nrows < AllocationTableConfig.DATA_START_ROW:
                continue
            
            # 读取该品番的数据
            product_data = self._read_product_sheet(sheet, sheet_name)
            if product_data:
                self.products_data.append(product_data)
        
        return {
            'metadata': self.metadata,
            'products': self.products_data
        }

    def _read_xlsx(self) -> Dict:
        """使用openpyxl读取.xlsx文件"""
        self.engine = 'openpyxl'
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            # 跳过空sheet或太小的sheet
            if sheet.max_row < AllocationTableConfig.DATA_START_ROW:
                continue
                
            # 读取该品番的数据
            product_data = self._read_product_sheet(sheet, sheet_name)
            if product_data:
                self.products_data.append(product_data)
                
        return {
            'metadata': self.metadata,
            'products': self.products_data
        }
    
    def _get_cell_value(self, sheet, row, col):
        """获取单元格值，屏蔽不同库的差异"""
        if self.engine == 'xlrd':
            try:
                if row >= sheet.nrows or col >= sheet.ncols:
                    return None
                return sheet.cell(row, col).value
            except:
                return None
        else: # openpyxl
            try:
                # openpyxl is 1-based
                return sheet.cell(row=row+1, column=col+1).value
            except:
                return None

    def _get_sheet_dims(self, sheet):
        """获取sheet的维度(rows, cols)"""
        if self.engine == 'xlrd':
            return sheet.nrows, sheet.ncols
        else:
            return sheet.max_row, sheet.max_column

    def _read_product_sheet(self, sheet, sheet_name: str) -> Dict:
        """
        读取单个品番sheet的数据
        
        Args:
            sheet: sheet对象
            sheet_name: sheet名称（品番）
            
        Returns:
            该品番的数据字典
        """
        try:
            # 提取元数据
            metadata = self._extract_metadata(sheet)
            
            # 读取表头（颜色和尺码信息）
            sku_columns = self._read_sku_columns(sheet)
            
            # 读取店铺数据
            stores_data = self._read_stores_data(sheet, sku_columns)
            
            return {
                'product_code': sheet_name,  # 品番
                'metadata': metadata,
                'sku_columns': sku_columns,   # SKU列信息
                'stores': stores_data         # 店铺数据
            }
            
        except Exception as e:
            print(f"警告: 读取sheet '{sheet_name}' 失败: {e}")
            return None
    
    def _extract_metadata(self, sheet) -> Dict:
        """提取元数据（管理No、納期等）"""
        metadata = {}
        nrows, ncols = self._get_sheet_dims(sheet)
        
        try:
            # 提取カンパニー
            if nrows > AllocationTableConfig.COMPANY_ROW:
                val = self._get_cell_value(sheet, 
                    AllocationTableConfig.COMPANY_ROW,
                    AllocationTableConfig.COMPANY_COL)
                if val:
                    metadata['company'] = val
            
            # 提取納期
            if nrows > AllocationTableConfig.DELIVERY_DATE_ROW:
                val = self._get_cell_value(sheet,
                    AllocationTableConfig.DELIVERY_DATE_ROW,
                    AllocationTableConfig.DELIVERY_DATE_COL)
                if val:
                    metadata['delivery_date'] = val
            
            # 提取品番（从元数据区域）
            if nrows > AllocationTableConfig.PRODUCT_CODE_ROW:
                val = self._get_cell_value(sheet,
                    AllocationTableConfig.PRODUCT_CODE_ROW,
                    AllocationTableConfig.PRODUCT_CODE_COL)
                if val:
                    metadata['product_code'] = val
            
            # 提取管理No（直接从已知位置读取）
            if (nrows > AllocationTableConfig.KANRI_NO_LABEL_ROW and 
                ncols > AllocationTableConfig.KANRI_NO_VALUE_COL):
                val = self._get_cell_value(sheet,
                    AllocationTableConfig.KANRI_NO_LABEL_ROW,
                    AllocationTableConfig.KANRI_NO_VALUE_COL)
                if val:
                    # 转换为字符串，如果是数字则转为整数字符串
                    if isinstance(val, float):
                        val = int(val)
                    metadata['kanri_no'] = str(val)
            
            # 提取店着日（直接从已知位置读取）
            if (nrows > AllocationTableConfig.STORE_DATE_LABEL_ROW and
                ncols > AllocationTableConfig.STORE_DATE_VALUE_COL):
                val = self._get_cell_value(sheet,
                    AllocationTableConfig.STORE_DATE_LABEL_ROW,
                    AllocationTableConfig.STORE_DATE_VALUE_COL)
                if val:
                    metadata['store_date'] = str(val)
            
        except Exception as e:
            print(f"警告: 提取元数据失败: {e}")
        
        # 使用第一个sheet的元数据作为全局元数据
        if not self.metadata:
            self.metadata = metadata.copy()
        
        return metadata
    
    def _read_sku_columns(self, sheet) -> List[Dict]:
        """
        读取SKU列信息（カラー和サイズ的组合）
        """
        sku_columns = []
        nrows, ncols = self._get_sheet_dims(sheet)
        
        # 读取カラー行（COL_FIRST_COLOR之后的列）
        color_row = AllocationTableConfig.HEADER_ROW
        size_row = AllocationTableConfig.SIZE_ROW
        
        # 从第一个颜色列开始遍历
        col_idx = AllocationTableConfig.COL_FIRST_COLOR
        current_color = None
        
        while col_idx < ncols:
            # 读取颜色
            color_val = self._get_cell_value(sheet, color_row, col_idx)
            if color_val:
                # 新的颜色值
                color_str = str(color_val).strip()
                # 过滤掉非数据列（如"カラー"标签列）
                if color_str and color_str != "カラー":
                    current_color = color_str
            
            # 读取尺码
            size_val = self._get_cell_value(sheet, size_row, col_idx)
            size_value = str(size_val).strip() if size_val else ""
            
            # 过滤掉无效的尺码值（如"サイズ"、"合計"等）
            invalid_sizes = ["サイズ", "合計", "合计", "小計", "Total"]
            is_valid_size = size_value and size_value not in invalid_sizes
            
            # 如果有颜色和有效尺码，记录这个SKU列
            if current_color and is_valid_size:
                sku_columns.append({
                    'column_index': col_idx,
                    'color': current_color,
                    'size': size_value
                })
            
            col_idx += 1
        
        return sku_columns
    
    def _read_stores_data(self, sheet, sku_columns: List[Dict]) -> List[Dict]:
        """
        读取店铺数据
        """
        stores_data = []
        nrows, ncols = self._get_sheet_dims(sheet)
        
        # 从数据开始行读取
        for row_idx in range(AllocationTableConfig.DATA_START_ROW, nrows):
            # 读取店铺基本信息
            no_val = self._get_cell_value(sheet, row_idx, AllocationTableConfig.COL_NO)
            store_code_val = self._get_cell_value(sheet, row_idx, AllocationTableConfig.COL_STORE_CODE)
            store_name_val = self._get_cell_value(sheet, row_idx, AllocationTableConfig.COL_STORE_NAME)
            type_val = self._get_cell_value(sheet, row_idx, AllocationTableConfig.COL_TYPE)
            
            # 如果没有店铺代码或名称，跳过这行（可能是汇总行）
            if not store_code_val or not store_name_val:
                continue
            
            # 读取该店铺的SKU配货数量
            sku_quantities = {}
            for sku_info in sku_columns:
                col_idx = sku_info['column_index']
                qty_val = self._get_cell_value(sheet, row_idx, col_idx)
                qty = 0
                if qty_val:
                    try:
                        qty = float(qty_val)
                    except:
                        qty = 0
                
                # 构建SKU key (color_size)
                sku_key = f"{sku_info['color']}_{sku_info['size']}"
                sku_quantities[sku_key] = qty
            
            # 只有当至少有一个SKU数量大于0时才记录这个店铺
            if any(qty > 0 for qty in sku_quantities.values()):
                store_data = {
                    'no': no_val if no_val else "",
                    'type': str(int(type_val)) if isinstance(type_val, (int, float)) and type_val else str(type_val) if type_val else "",
                    'store_code': str(int(store_code_val)) if isinstance(store_code_val, float) else str(store_code_val),
                    'store_name': str(store_name_val),
                    'sku_quantities': sku_quantities
                }
                
                # 读取ランク（如果有）
                if AllocationTableConfig.COL_RANK < ncols:
                    rank_val = self._get_cell_value(sheet, row_idx, AllocationTableConfig.COL_RANK)
                    if rank_val:
                        store_data['rank'] = str(rank_val)
                
                stores_data.append(store_data)
        
        return stores_data
    
    def close(self):
        """关闭工作簿"""
        if self.workbook:
            if self.engine == 'xlrd':
                self.workbook.release_resources()
            elif self.engine == 'openpyxl':
                self.workbook.close()
            self.workbook = None
