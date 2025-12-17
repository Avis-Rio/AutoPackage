#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将.xls模板转换为.xlsx格式
"""
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def convert_xls_to_xlsx(xls_path, xlsx_path):
    """将xls文件转换为xlsx"""
    print(f"读取 {xls_path}...")
    rb = xlrd.open_workbook(xls_path, formatting_info=True)
    
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认sheet
    
    # 转换每个sheet
    for sheet_idx in range(rb.nsheets):
        rs = rb.sheet_by_index(sheet_idx)
        ws = wb.create_sheet(rs.name)
        
        print(f"转换 sheet: {rs.name} ({rs.nrows} 行 x {rs.ncols} 列)")
        
        # 复制数据和基本样式
        for row_idx in range(rs.nrows):
            for col_idx in range(rs.ncols):
                cell = rs.cell(row_idx, col_idx)
                ws_cell = ws.cell(row=row_idx+1, column=col_idx+1)
                
                # 复制值
                ws_cell.value = cell.value
                
                # 尝试复制样式
                try:
                    if hasattr(cell, 'xf_index') and cell.xf_index is not None:
                        xf = rb.xf_list[cell.xf_index]
                        
                        # 字体
                        font_obj = rb.font_list[xf.font_index]
                        ws_cell.font = Font(
                            name=font_obj.name.decode('latin1') if isinstance(font_obj.name, bytes) else font_obj.name,
                            size=font_obj.height / 20.0,
                            bold=font_obj.bold,
                            italic=font_obj.italic
                        )
                        
                        # 对齐
                        ws_cell.alignment = Alignment(
                            horizontal='center' if xf.alignment.hor_align == 2 else 'left',
                            vertical='center' if xf.alignment.vert_align == 1 else 'top',
                            wrap_text=xf.alignment.text_wrapped
                        )
                        
                        # 背景色
                        bg = xf.background
                        if bg.pattern_colour_index != 64:  # 64 = 无背景
                            # xlrd颜色索引到RGB的简单映射
                            color_map = {
                                13: 'FFFF00',  # 黄色
                                22: 'D3D3D3',  # 灰色
                            }
                            rgb = color_map.get(bg.pattern_colour_index, 'FFFFFF')
                            ws_cell.fill = PatternFill(start_color=rgb, end_color=rgb, fill_type='solid')
                        
                        # 边框
                        border_style = 'thin'
                        ws_cell.border = Border(
                            left=Side(style=border_style) if xf.border.left_line_style else None,
                            right=Side(style=border_style) if xf.border.right_line_style else None,
                            top=Side(style=border_style) if xf.border.top_line_style else None,
                            bottom=Side(style=border_style) if xf.border.bottom_line_style else None
                        )
                except Exception as e:
                    pass
    
    print(f"保存为 {xlsx_path}...")
    wb.save(xlsx_path)
    print("转换完成！")

if __name__ == "__main__":
    convert_xls_to_xlsx("template.xls", "template.xlsx")
